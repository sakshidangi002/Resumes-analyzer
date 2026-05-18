"""Reports: attendance, leave, headcount, salary summary; export Excel/PDF."""
from io import BytesIO
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.db.session import get_db
from app.models import (
    User,
    AttendanceRecord,
    LeaveRequest,
    Employee,
    Payslip,
    PayrollPeriod,
    Department,
    LeaveAllocation,
)
from app.api.deps import get_current_user, require_roles
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from calendar import month_abbr, monthrange
from sqlalchemy.orm import joinedload
from datetime import date, timedelta

router = APIRouter()


@router.get("/attendance/monthly")
def monthly_attendance_report(
    month: int = Query(...),
    year: int = Query(...),
    department_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["Admin", "HR", "Manager"])),
):
    actual_days = 30
    start = date(year, month, 1)
    if month == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, month + 1, 1)
        from datetime import timedelta
        end = end - timedelta(days=1)
    # Get all employees in the department or all active employees
    emp_query = db.query(Employee).filter(Employee.employment_status == "Active")
    if department_id is not None:
        emp_query = emp_query.filter(Employee.department_id == department_id)
    employees = emp_query.all()
    emp_ids = [e.id for e in employees]

    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id.in_(emp_ids),
        AttendanceRecord.date >= start,
        AttendanceRecord.date <= end,
    ).all()
    
    leave_reqs = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id.in_(emp_ids),
        LeaveRequest.status == "APPROVED",
        LeaveRequest.start_date <= end,
        LeaveRequest.end_date >= start,
    ).all()

    # Determine the end of the calculation range (either end of month or today)
    today = date.today()
    if today.year == year and today.month == month:
        calc_end = today
    else:
        calc_end = end

    # Calculate weekend days (Saturdays and Sundays) up to calc_end
    total_weekends = 0
    curr = start
    while curr <= calc_end:
        if curr.weekday() >= 5: # 5=Saturday, 6=Sunday
            total_weekends += 1
        curr = curr + timedelta(days=1)

    summary = {}
    for emp in employees:
        summary[emp.id] = {
            "employee_code": emp.employee_code,
            "first_name": emp.first_name,
            "last_name": emp.last_name,
            "present": 0,
            "absent": 0,
            "half_day": 0,
            "on_leave": 0,
            "week_off": total_weekends,
            "total_attendance": float(total_weekends + (calc_end - start).days + 1 - total_weekends), # Start with days passed
            "total_leaves": 0,
            "working_days": actual_days
        }
    
    # Reset total_attendance to 0 and build it up from actual status up to today
    for emp_id in summary:
        summary[emp_id]["total_attendance"] = float(summary[emp_id]["week_off"])

    for rec in records:
        if rec.date > calc_end: continue
        s = summary.get(rec.employee_id)
        if not s: continue
        if rec.status in ("PRESENT", "SHORT"):
            s["present"] += 1
            s["total_attendance"] += 1
        elif rec.status == "HALF_DAY":
            s["half_day"] += 1
            s["total_attendance"] += 0.5
        elif rec.status == "ABSENT":
            s["absent"] += 1
            # s["total_attendance"] remains same (no addition)
        elif rec.status in ("PAID_LEAVE", "ON_LEAVE"):
            # Check if it was already handled by leave_reqs to avoid double counting
            # Wait, leave_reqs loop happens AFTER this. So if we process ON_LEAVE here, 
            # and leave_reqs processes it too, it will double count.
            # However, if it's PAID_LEAVE (manual), it won't be in leave_reqs.
            if rec.status == "PAID_LEAVE":
                s["total_leaves"] += 1
                s["on_leave"] += 1
                s["total_attendance"] += 1
        elif rec.status in ("WEEKLY_OFF", "HOLIDAY"):
            # Weekends are already in the base total_weekends
            # If it's a weekday holiday, add it
            if rec.status == "HOLIDAY" and rec.date.weekday() < 5:
                s["total_attendance"] += 1

    for req in leave_reqs:
        s = summary.get(req.employee_id)
        if not s: continue
        o_start = max(req.start_date, start)
        o_end = min(req.end_date, calc_end) # Only up to calc_end
        if o_start <= o_end:
            overlap_days = (o_end - o_start).days + 1
            leaves_to_add = (0.5 if req.is_half_day else 1.0) * overlap_days
            s["total_leaves"] += leaves_to_add
            s["on_leave"] += leaves_to_add
            s["total_attendance"] += leaves_to_add # Add leaves since they are part of attendance

    # Convert to list and sort by employee_code ascending
    summary_list = [{"employee_id": eid, **vals} for eid, vals in summary.items()]
    summary_list.sort(key=lambda x: str(x.get("employee_code", "")))

    return {"month": month, "year": year, "summary": summary_list}


@router.get("/leave/usage")
def leave_usage_report(
    financial_year_id: int | None = Query(None),
    department_id: int | None = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["Admin", "HR", "Manager"])),
):
    q = db.query(LeaveAllocation).join(Employee, Employee.id == LeaveAllocation.employee_id)
    if financial_year_id:
        q = q.filter(LeaveAllocation.financial_year_id == financial_year_id)
    if department_id:
        q = q.filter(Employee.department_id == department_id)
    allocs = q.all()
    return {"allocations": [{"employee_id": a.employee_id, "leave_type_id": a.leave_type_id, "allocated_days": float(a.allocated_days), "used_days": float(a.used_days)} for a in allocs]}


@router.get("/headcount/department")
def department_headcount(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["Admin", "HR", "Manager"])),
):
    q = db.query(Employee.department_id, func.count(Employee.id)).filter(
        Employee.employment_status == "Active"
    ).group_by(Employee.department_id)
    rows = q.all()
    dept_names = {d.id: d.name for d in db.query(Department).all()}
    return {"headcount": [{"department_id": r[0], "department_name": dept_names.get(r[0], ""), "count": r[1]} for r in rows]}


@router.get("/salary/summary")
def salary_summary(
    year: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["Admin", "HR"])),
):
    q = db.query(PayrollPeriod.month, func.sum(Payslip.net_salary), func.count(Payslip.id)).join(
        Payslip, Payslip.payroll_period_id == PayrollPeriod.id
    ).filter(PayrollPeriod.year == year).group_by(PayrollPeriod.month)
    rows = q.all()
    return {"year": year, "monthly": [{"month": r[0], "total_net": float(r[1]), "employee_count": r[2]} for r in rows]}


@router.get("/export/attendance-excel")
def export_attendance_excel(
    month: int = Query(...),
    year: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(["Admin", "HR", "Manager"])),
):
    from calendar import month_name, monthrange
    
    employees = db.query(Employee).filter(
        Employee.employment_status == "Active"
    ).order_by(Employee.employee_code).all()
    
    start = date(year, month, 1)
    _, num_days = monthrange(year, month)
    end = date(year, month, num_days)
    
    # Range for current month logic
    today = date.today()
    calc_end = today if (today.year == year and today.month == month) else end

    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.date >= start,
        AttendanceRecord.date <= end,
    ).all()
    
    leave_reqs = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id.in_([e.id for e in employees]),
        LeaveRequest.status == "APPROVED",
        LeaveRequest.start_date <= end,
        LeaveRequest.end_date >= start,
    ).all()
    
    attendance_map = {}
    for r in records:
        attendance_map[(r.employee_id, r.date)] = r

    leave_map = {}
    for req in leave_reqs:
        o_start = max(req.start_date, start)
        o_end = min(req.end_date, end)
        curr = o_start
        while curr <= o_end:
            leave_map[(req.employee_id, curr)] = 0.5 if req.is_half_day else 1.0
            curr += timedelta(days=1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Styles
    side = Side(style='thin', color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    center_align = Alignment(horizontal='center', vertical='center')
    bold_font = Font(bold=True)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Header Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_days + 7)
    title = ws.cell(row=1, column=1, value=f"ATTENDANCE REPORT - {month_name[month].upper()} {year}")
    title.font = Font(size=16, bold=True)
    title.alignment = center_align

    # Header Row 2: Labels
    headers = ["No.", "Code", "Name"] + [str(d) for d in range(1, num_days + 1)] + ["Present", "Absent", "Leave", "Total"]
    for col, val in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
        
        # Merge vertical for No, Code, Name, and Summary columns
        if col <= 3 or col > num_days + 3:
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    # Header Row 3: Weekdays
    day_labels = ["M", "T", "W", "TH", "F", "S", "SU"]
    for d in range(1, num_days + 1):
        curr_date = date(year, month, d)
        day_name = day_labels[curr_date.weekday()]
        cell = ws.cell(row=3, column=d + 3, value=day_name)
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Column Widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    for i in range(4, num_days + 4):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5

    # Data Rows
    for idx, emp in enumerate(employees, start=1):
        row_idx = idx + 3
        ws.cell(row=row_idx, column=1, value=idx).border = border
        ws.cell(row=row_idx, column=1).alignment = center_align
        ws.cell(row=row_idx, column=2, value=emp.employee_code).border = border
        ws.cell(row=row_idx, column=2).alignment = center_align
        ws.cell(row=row_idx, column=3, value=emp.full_name).border = border
        
        present_count = 0
        absent_count = 0
        leave_count = 0
        total_weekends = 0
        
        # Calculate weekends up to today/end
        curr_test = start
        while curr_test <= calc_end:
            if curr_test.weekday() >= 5: total_weekends += 1
            curr_test += timedelta(days=1)
            
        total_attendance = 30.0 # Start with 30 as per request
        
        for d in range(1, num_days + 1):
            curr_date = date(year, month, d)
            col_idx = d + 3
            val = ""
            
            # Weekend Check
            is_weekend = curr_date.weekday() >= 5
            if is_weekend: val = "WO"
            
            # Record Check
            rec = attendance_map.get((emp.id, curr_date))
            if rec:
                if rec.status in ("PRESENT", "SHORT"):
                    val = "P"
                    if curr_date <= calc_end: present_count += 1
                elif rec.status == "HALF_DAY":
                    val = "HD"
                    if curr_date <= calc_end:
                        present_count += 0.5
                        total_attendance -= 0.5
                elif rec.status == "ABSENT":
                    val = "A"
                    if curr_date <= calc_end:
                        absent_count += 1
                        total_attendance -= 1
                elif rec.status == "PAID_LEAVE":
                    val = "L"
                    if curr_date <= calc_end:
                        leave_count += 1
                        total_attendance -= 1
            
            # Leave Check
            l_val = leave_map.get((emp.id, curr_date))
            if l_val:
                val = "L" if l_val == 1.0 else "HL"
                if curr_date <= calc_end:
                    leave_count += l_val
                    total_attendance -= l_val
            
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = center_align
            if is_weekend:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # Summary Columns
        ws.cell(row=row_idx, column=num_days + 4, value=present_count).border = border
        ws.cell(row=row_idx, column=num_days + 5, value=absent_count).border = border
        ws.cell(row=row_idx, column=num_days + 6, value=leave_count).border = border
        
        # Final Total Attendance capped at 30 logic
        # But if it's mid-month, we should probably show the projected or current?
        # User said "Always 30" and "Minus not add". 
        # So it should be 30 - (Absents so far) - (Leaves so far)
        final_total = 30.0 - absent_count - leave_count
        
        total_cell = ws.cell(row=row_idx, column=num_days + 7, value=final_total)
        total_cell.border = border
        total_cell.font = bold_font
        total_cell.alignment = center_align

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=attendance_report_{year}_{month}.xlsx"}
    )


    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": f"attachment; filename=attendance_sheet_{year}_{month}.xlsx"}
    )
